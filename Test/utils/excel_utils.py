import openpyxl
from openpyxl.styles import Alignment


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno, colno).value


def readData_without_open_file(sheet, rowno, colno):
    return sheet.cell(rowno, colno).value


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


def readsinglerow(file, sheetName, rownum, colnum):
    single_row = []
    for i in range(1, colnum + 1):
        data = readData(file, sheetName, rownum, i)
        single_row.append(data)
    return single_row


def readsinglecol(file, sheetName, srow, maxrownum, colnum):
    single_col = []
    for i in range(srow, maxrownum + 1):
        data = readData(file, sheetName, i, colnum)
        single_col.append(data)
    return single_col


def readsinglecol2(file, sheetName, srow, maxrownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    single_col = []
    for i in range(srow, maxrownum + 1):
        data = readData_without_open_file(sheet, i, colnum)
        single_col.append(data)
    return single_col


def writelistoflist(file, sheetName, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    row = 1
    for element in data:
        col = 1
        for d in element:
            if d:
                sheet.cell(row, col).value = d
            col += 1
        row += 1
    wb.save(file)


def writesinglerow(file, sheetName, rownum, colnum, scol, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    for i in range(1, colnum + 1):
        sheet.cell(rownum, i + scol).value = data[i - 1]
        # print(data[i - 1])
    wb.save(file)


def writesinglecol(file, sheetName, rownum, colnum, srow, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    for i in range(1, rownum + 1):
        sheet.cell(i + srow, colnum).value = str(data[i - 1])
        # print(data[i - 1])
    wb.save(file)


def mergecell(file, sheetName, mcell1, mcell2, rownum, colnum, ):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.merge_cells(f'{mcell1}:{mcell2}')
    cell = sheet.cell(rownum, colnum)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(file)


def writeinmergecell(file, sheetName, mcell1, mcell2, rownum, colnum, scol, data):
    mergecell(file, sheetName, mcell1, mcell2, rownum, scol + 1)
    writesinglerow(file, sheetName, rownum, colnum, scol, data)


def remove_items(list1, item):
    res = [i for i in list1 if i != item]
    return res


# readonly

def get_row_count_read_only(file, sheet_name):
    workbook = openpyxl.load_workbook(filename=file, read_only=True)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value)
    # max_row = sheet.max_row
    # print(max_row)
    workbook.close()
    # return max_row


def get_col_count_read_only(file, sheet_name):
    workbook = openpyxl.load_workbook(file, read_only=True)
    sheet = workbook[sheet_name]
    max_column = sheet.max_column
    workbook.close()
    return max_column


def read_single_col_read_only(file, sheet_name, starting_row, max_row_num, col_num):
    single_col = []
    for i in range(starting_row, max_row_num + 1):
        data = readData(file, sheet_name, i, col_num)
        single_col.append(data)
    return single_col
